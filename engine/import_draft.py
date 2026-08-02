from pathlib import Path
import json

from openpyxl import load_workbook

from players.service import load_players


# =========================
# CONSTANTS
# =========================

MANAGERS = [
    "matej",
    "goran",
    "johnny",
    "kuba",
    "paulie",
    "francesco"
]

MANAGER_COLUMNS = {
    "matej": 2,
    "goran": 4,
    "johnny": 6,
    "kuba": 8,
    "paulie": 10,
    "francesco": 12
}

ROLE_ROWS = {
    "P": (6, 8),      # Portieri
    "D": (9, 16),     # Difensori
    "C": (17, 24),    # Centrocampisti
    "A": (25, 30)     # Attaccanti
}

BUDGET_ROW = 32
COACH_ROW = 34


# =========================
# HELPERS
# =========================

def find_player(players, name):

    name = str(name).strip().lower()

    for player in players:

        if player["name"].strip().lower() == name:
            return player.copy()

    print(f"Nenalezen hráč: {name}")
    return None


def save_squad(manager, squad):

    path = Path(f"data/{manager}/squad.json")

    data = {
        "manager": manager.capitalize(),
        "budget": squad["budget"],
        "coach": squad["coach"],
        "squad": squad["players"]
    }

    with open(
        path,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            data,
            f,
            indent=4,
            ensure_ascii=False
        )


# =========================
# BUILD SQUAD
# =========================

def build_squad(sheet, manager, players):

    column = MANAGER_COLUMNS[manager]

    squad = {
        "coach": sheet.cell(COACH_ROW, column).value,
        "budget": sheet.cell(BUDGET_ROW, column + 1).value,
        "players": []
    }

    for role, rows in ROLE_ROWS.items():

        start_row, end_row = rows

        for row in range(start_row, end_row + 1):

            player_name = sheet.cell(row, column).value
            buy_price = sheet.cell(row, column + 1).value

            if not player_name:
                continue

            player = find_player(
                players,
                player_name
            )

            if player is None:

                raise Exception(
                    f"Hráč '{player_name}' nebyl nalezen v players.json."
                )

            player["buy_price"] = buy_price

            squad["players"].append(
                player
            )

    return squad


# =========================
# IMPORT DRAFT
# =========================

def import_draft(excel_path):

    workbook = load_workbook(
        excel_path,
        data_only=True
    )

    sheet = workbook["Draft"]

    players = load_players()

    report = []

    for manager in MANAGERS:

        squad = build_squad(
            sheet,
            manager,
            players
        )

        if len(squad["players"]) != 25:

            raise Exception(
                f"{manager} má {len(squad['players'])} hráčů místo 25."
            )

        save_squad(
            manager,
            squad
        )

        report.append({
            "manager": manager.capitalize(),
            "players": len(squad["players"]),
            "coach": squad["coach"],
            "budget": squad["budget"]
        })

    return report


if __name__ == "__main__":

    report = import_draft("Draft 2025-26.xlsx")

    print()

    print("===== IMPORT REPORT =====")

    for row in report:

        print(
            f"{row['manager']}: "
            f"{row['players']} hráčů | "
            f"{row['coach']} | "
            f"{row['budget']}M"
        )