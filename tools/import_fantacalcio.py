import json
from pathlib import Path

from openpyxl import load_workbook


# =========================
# FILES
# =========================

PROJECT_DIR = Path(__file__).resolve().parent.parent

IMPORT_DIR = PROJECT_DIR / "imports"

excel_files = sorted(
    IMPORT_DIR.glob("Quotazioni*.xlsx")
)

if not excel_files:

    raise FileNotFoundError(
        "\nVe složce 'imports' nebyl nalezen žádný "
        "soubor Quotazioni*.xlsx"
    )

SOURCE_FILE = excel_files[0]

OUTPUT_FILE = PROJECT_DIR / "data" / "players.json"


# =========================
# ROLE CONVERSION
# =========================

ROLE_MAP = {
    "P": "P",
    "D": "D",
    "C": "C",
    "A": "A"
}


# =========================
# LOAD EXCEL
# =========================

workbook = load_workbook(
    SOURCE_FILE,
    data_only=True
)

# =========================
# SHEETS
# =========================

sheets = [

    (workbook["Tutti"], True),

    (workbook["Ceduti"], False)

]


# =========================
# FIND COLUMNS
# =========================

headers = {}

HEADER_ROW = 2

header_sheet = workbook["Tutti"]

for column in range(1, header_sheet.max_column + 1):

    value = header_sheet.cell(
        row=HEADER_ROW,
        column=column
    ).value

    if value is not None:

        headers[
            str(value).strip()
        ] = column


players = []

used_ids = set()


# =========================
# READ PLAYERS
# =========================

for sheet, active in sheets:

    for row in range(HEADER_ROW + 1, sheet.max_row + 1):

        player_id = sheet.cell(
            row=row,
            column=headers["Id"]
        ).value

        if player_id is None:
            continue

        player_id = int(player_id)

        if player_id in used_ids:
            continue

        used_ids.add(player_id)

        name = sheet.cell(
            row=row,
            column=headers["Nome"]
        ).value

        team = sheet.cell(
            row=row,
            column=headers["Squadra"]
        ).value

        role = sheet.cell(
            row=row,
            column=headers["R"]
        ).value

        qi = sheet.cell(
            row=row,
            column=headers["Qt.I"]
        ).value

        qa = sheet.cell(
            row=row,
            column=headers["Qt.A"]
        ).value

        players.append({

            "id": player_id,

            "name": str(name).strip(),

            "team": str(team).strip(),

            "role": ROLE_MAP.get(
                str(role).strip(),
                str(role).strip()
            ),

            "qi": int(qi),

            "qa": int(qa),

            "active": active

        })


# =========================
# SORT PLAYERS
# =========================

ROLE_ORDER = {
    "P": 0,
    "D": 1,
    "C": 2,
    "A": 3
}

players.sort(
    key=lambda p: (
        ROLE_ORDER[p["role"]],
        p["team"],
        p["name"]
    )
)


# =========================
# SAVE JSON
# =========================

OUTPUT_FILE.parent.mkdir(
    exist_ok=True
)

with open(
    OUTPUT_FILE,
    "w",
    encoding="utf-8"
) as f:

    json.dump(
        players,
        f,
        ensure_ascii=False,
        indent=4
    )


print()
print("=" * 40)
print("IMPORT HOTOV")
print("=" * 40)
print()
print(f"Importováno hráčů: {len(players)}")
print(f"Soubor uložen jako: {OUTPUT_FILE}")
print()